#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each county.

import time, openpyxl, pprint, os
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

start_time = time.time()
print('Opening workbook...')
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet = wb.active
number = input('Please put your number: ')

print('Creating rows and columns...')
for rowNum in range(1,int(number)+1):
	sheet.cell(row=rowNum, column=1).value = rowNum
for colNum in range(1,int(number)):
	sheet.cell(column=colNum, row=1).value = colNum

for i in range(1,int(number)+1):
	for j in range(1,int(number)+1):
		sheet.cell(row=i, column=j).value = i*j

fileName = "multiplicationExcelTable.xlsx"
wb.save(fileName)

try:
	print('Oppening excel file...')
	file = os.startfile(fileName)
except Exception as exc:
	print('There was a problem: %s' % (exc))

end_time = time.time()
print("\n--- %s seconds ---" % round((end_time - start_time),2))
